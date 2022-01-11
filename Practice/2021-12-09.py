# 여러 엑셀에서 데이터 불러오기
import os
from openpyxl import load_workbook

path = "C:/...."
file_list = os.listdir(path)
print(file_list)

results = []

for file_name_raw in file_list:
    """Basic Information Extract"""

    file_name = path + '/' + file_name_raw
    wb = load_workbook(filename = file_name, data_only=True)
    ws = wb['基本情報'] # assign specific sheet

    basicinfo = []
    basicinfo.append(file_name_raw)
    basicinfo.append(ws['B3'].value)
    basicinfo.append(ws['B4'].value)
    basicinfo.append(ws['B6'].value)
    basicinfo.append(ws['D1'].value)

    results.append(basicinfo)

print(results)

for i in results:
    ws.append(i)

# wb.save("C:/..../tmp.xlsx")
wb.save("C:/..../Basic_info.xlsx")